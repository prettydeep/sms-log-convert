import argparse
import xml.etree.ElementTree as ET
from openpyxl import Workbook

parser = argparse.ArgumentParser()
parser.add_argument("--input_file", type=str, required=True, help="Path to call-log xml data file.")
parser.add_argument("--output_file", type=str, required=True, help="Path to output xlsx file.")
args = parser.parse_args()

input_file = args.input_file
output_file = args.output_file

tree = ET.parse(input_file)
root = tree.getroot()

workbook = Workbook()
sheet = workbook.active
sheet.title = "Call Log"

header = ['Call Number', 'Duration', 'Date', 'Type', 'Readable Date', 'Contact Name']
sheet.append(header)

call_type_mapping = {
    "1": "Incoming",
    "2": "Outgoing",
    "3": "Missed",
    "5": "Declined",
    "6": "Rejected/Spam"
}

for elem in root:
    if elem.tag == 'call':
        current_row = len(sheet['A']) + 1

        sheet.cell(row=current_row, column=1).value = elem.attrib.get('number')
        sheet.cell(row=current_row, column=2).value = elem.attrib.get('duration')
        sheet.cell(row=current_row, column=3).value = elem.attrib.get('date')
        
        call_type = call_type_mapping.get(elem.attrib.get('type'), "Unknown")
        sheet.cell(row=current_row, column=4).value = call_type
        
        sheet.cell(row=current_row, column=5).value = elem.attrib.get('readable_date')
        sheet.cell(row=current_row, column=6).value = elem.attrib.get('contact_name')

workbook.save(output_file)
