import base64
import mimetypes
import argparse
import tempfile
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.drawing.image import Image

parser = argparse.ArgumentParser()
parser.add_argument("--input_file", type=str, required=True, help="Path to SMS/MMS xml data file.")
parser.add_argument("--output_file", type=str, required=True, help="Path to output xlsx file.")
args = parser.parse_args()

input_file = args.input_file
output_file = args.output_file

mimetypes.add_type('image/mpo', '.mpo')

tree = ET.parse(input_file)
root = tree.getroot()

workbook = Workbook()
sheet = workbook.active
sheet.title = "SMS and MMS Log"

header = ['Protocol', 'Date', 'Address', 'Type', 'Body', 'Contact Name', 'Image']
sheet.append(header)

def process_image_data(image_data):
    with tempfile.NamedTemporaryFile(suffix=".jpg", delete=False) as temp_file:
        temp_file.write(base64.b64decode(image_data))
        return temp_file.name

for elem in root:
    if elem.tag == 'sms' or elem.tag == 'mms':
        current_row = len(sheet['A']) + 1

        sheet.cell(row=current_row, column=1).value = 'SMS' if elem.tag == 'sms' else 'MMS'
        sheet.cell(row=current_row, column=2).value = elem.attrib.get('readable_date')
        sheet.cell(row=current_row, column=3).value = elem.attrib.get('address')
        
        if elem.tag == 'sms':
            msg_type = 'Received' if elem.attrib.get('type') == '1' else 'Sent'
        elif elem.tag == 'mms':
            msg_type = 'Received' if elem.attrib.get('msg_box') == '1' else 'Sent'
        sheet.cell(row=current_row, column=4).value = msg_type
        
        if elem.tag == 'sms':
            body_text = elem.attrib.get('body')
        elif elem.tag == 'mms':
            part_elem = elem.find(".//part[@ct='text/plain']")
            body_text = part_elem.attrib.get("text") if part_elem is not None else None
        sheet.cell(row=current_row, column=5).value = body_text
        
        sheet.cell(row=current_row, column=6).value = elem.attrib.get('contact_name')
        
        if elem.tag == 'mms':
            image_elems = [e for e in elem.findall(".//part") if e.attrib['ct'].startswith('image/')]
            if image_elems:
                try:
                    image_elem = image_elems[0]
                    image_data = image_elem.attrib['data']
                    temp_image_file = process_image_data(image_data)
                    img = Image(temp_image_file)
                    sheet.column_dimensions['G'].width = img.width // 6
                    sheet.row_dimensions[current_row].height = img.height
                    sheet.add_image(img, f'G{current_row}')
                except:
                    print(f'file {temp_image_file} is not a valid image file ')

workbook.save(output_file)
